# Job Tracker Auto-Decline Script
# Scans Outlook inbox for rejection emails and updates job_tracker.xlsx
#
# Requirements:
#     pip install openpyxl pywin32
#     python -m pywin32_postinstall -install
#
# Usage:
#     python job_tracker_auto_decline.py
#     python job_tracker_auto_decline.py --dry-run
#     python job_tracker_auto_decline.py --debug

import argparse
import re
import sys
from pathlib import Path
from datetime import datetime

try:
    import win32com.client
except ImportError:
    sys.exit("ERROR: pywin32 is not installed.\nRun:  pip install pywin32\nThen: python -m pywin32_postinstall -install")

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    sys.exit("ERROR: openpyxl is not installed.\nRun:  pip install openpyxl")


# ---------- Configuration ----------

DEFAULT_EXCEL = Path(__file__).parent / "job_tracker.xlsx"
COL_COMPANY    = 1   # Column A
COL_STATUS     = 4   # Column D
# --- Rejection phrases ---
# Covers: formal rejections, ATS auto-emails, polite declines,
# post-interview rejections, and non-English influenced phrasing
REJECTION_PHRASES = [
    # Moving forward / proceeding
    "we will not be moving forward",
    "we won't be moving forward",
    "we are not moving forward",
    "we're not moving forward",
    "not moving forward with your application",
    "not moving forward with you",
    "decided not to move forward",
    "have decided not to move forward",
    "unable to move forward",
    "we will not be proceeding",
    "we won't be proceeding",
    "we are not proceeding",
    "not proceeding with your application",
    "not proceeding with your candidacy",
    "unable to proceed with your candidacy",
    "unable to proceed with your application",
    "we're unable to proceed",
    "we are unable to proceed",
    # Other candidates chosen
    "we have decided to move forward with other candidates",
    "decided to move forward with other candidates",
    "moving forward with other candidates",
    "moving forward with another candidate",
    "we have chosen to move forward with another candidate",
    "we have chosen to move forward with other candidates",
    "we have chosen other candidates",
    "chosen to pursue other candidates",
    "we will be moving forward with other applicants",
    "pursuing other candidates",
    "pursuing other applicants",
    "pursuing other profiles",
    # Not selected
    "you have not been selected",
    "not been selected for",
    "not selected for this position",
    "not selected for this role",
    "not selected for the position",
    "not selected for the role",
    "not selected to move forward",
    "was not selected",
    "were not selected",
    # Application outcome
    "your application was not successful",
    "your application has been unsuccessful",
    "unsuccessful in your application",
    "we are unable to offer you",
    "we will not be offering you",
    "we won't be offering you",
    "no offer will be extended",
    "we will not be extending an offer",
    # Regret / sorry phrases
    "we regret to inform you",
    "we regret to let you know",
    "we regret to advise you",
    "we're sorry to inform you",
    "we are sorry to inform you",
    "we're sorry to let you know",
    "we are sorry to let you know",
    "we're sorry to advise you",
    "sorry to let you know",
    "sorry to inform you",
    # After careful consideration
    "after careful consideration",
    "after thorough consideration",
    "after reviewing your application",
    "after much consideration",
    "following careful consideration",
    "following our review",
    "having reviewed your application",
    # Position filled / closed
    "position has been filled",
    "role has been filled",
    "vacancy has been filled",
    "the position is no longer available",
    "we have filled the position",
    "we have closed this position",
    # Does not meet / fit
    "does not meet our requirements",
    "do not meet the requirements",
    "does not match our current needs",
    "not the right fit",
    "not a strong enough match",
    "not quite what we are looking for",
    "not quite what we're looking for",
    "your profile does not match",
    "your background does not match",
    "your experience does not match",
    # Generic decline
    "thank you for your interest, however",
    "thank you for applying, however",
    "appreciate your interest, however",
    "we have decided not to",
    "decided to decline",
    "we will not be considering",
    "no longer being considered",
    "your application will not be taken further",
    "will not be taken further",
    "will not be progressed",
    "your candidacy will not",
]

# --- Interview / invitation phrases ---
# Covers: phone screens, video calls, technical assessments,
# on-site interviews, recruiter reach-outs and ATS invites
INTERVIEW_PHRASES = [
    # Direct interview invite
    "invite you to an interview",
    "invited you to an interview",
    "invitation to interview",
    "invitation for an interview",
    "you are invited to interview",
    "you have been invited to interview",
    "we would like to invite you to interview",
    "pleased to invite you to an interview",
    "happy to invite you to an interview",
    "we would like to interview you",
    "we'd like to interview you",
    "like to interview you",
    "we want to interview you",
    # Selected / shortlisted
    "you have been selected for an interview",
    "you've been selected for an interview",
    "selected to move forward to an interview",
    "shortlisted for an interview",
    "you have been shortlisted",
    "you've been shortlisted",
    "moving forward with you to the interview",
    "moving you forward to the next stage",
    "moving you to the next round",
    "you have been selected for the next stage",
    "selected for the next round",
    "you have been selected to proceed",
    "you've been selected to proceed",
    # Schedule / set up
    "schedule an interview",
    "schedule a interview",
    "schedule your interview",
    "scheduling an interview",
    "like to schedule a time",
    "like to set up a time",
    "like to set up a call",
    "like to schedule a call",
    "like to arrange a call",
    "like to arrange an interview",
    "like to arrange a meeting",
    "schedule a meeting with you",
    "schedule time to talk",
    "schedule a conversation",
    "schedule a chat",
    "book an interview",
    "book a time with",
    "book a call",
    "book a meeting",
    # Next step / stage
    "next step is an interview",
    "next steps involve an interview",
    "next step in our process",
    "next stage of our process",
    "next stage of the recruitment",
    "next stage of the hiring",
    "move you to the next step",
    "progress to the next stage",
    "progress to the interview stage",
    "advance to the next round",
    "advance you to the interview",
    # Phone / video screen
    "phone screen",
    "phone interview",
    "video interview",
    "video call with",
    "virtual interview",
    "zoom interview",
    "teams interview",
    "google meet interview",
    "introductory call",
    "initial call",
    "screening call",
    "discovery call",
    "chat with our team",
    "chat with the team",
    "speak with our recruiter",
    "speak with the hiring manager",
    "talk with our team",
    # Technical assessment / test
    "technical assessment",
    "technical interview",
    "technical test",
    "coding assessment",
    "coding challenge",
    "coding test",
    "take-home assignment",
    "take home assignment",
    "complete an assessment",
    "complete a test",
    "online assessment",
    # Would like to meet / connect
    "would like to meet with you",
    "we would like to invite you",
    "we'd like to invite you",
    "we'd love to meet you",
    "would love to meet you",
    "would love to chat with you",
    "would love to connect",
    "pleased to meet with you",
    "happy to meet with you",
    "keen to learn more about you",
    "excited to learn more about you",
    # Recruiter reach out
    "we came across your profile",
    "we found your profile",
    "your profile caught our attention",
    "your background caught our attention",
    "your experience caught our attention",
    "i came across your profile",
    "i found your profile",
    "reach out regarding an opportunity",
    "reaching out about an opportunity",
    "reaching out about a role",
    "reaching out about a position",
    "open to exploring",
    "open to new opportunities",
    "interested in hearing more",
]


# --- Application confirmation phrases ---
APPLICATION_PHRASES = [
    "thank you for your application",
    "thank you for applying",
    "we have received your application",
    "we received your application",
    "your application has been received",
    "application has been submitted",
    "thank you for considering",
    "we are delighted that you",
    "we're delighted that you",
    "thanks for applying",
    "your application to",
    "we will be in touch",
    "we'll be in touch",
    "will get back to you",
    "will be in touch as soon as",
]

SHEET_NAME     = "Job Tracker"
DATA_START_ROW = 5   # Rows 1-4 are title/summary/blank/header
COL_ROLE  = 2   # Column B
COL_DATE  = 3   # Column C
COL_NOTES = 5   # Column E


# ---------- Outlook helpers ----------

# Set this to the email address of the Outlook account you want to scan.
# Example: "yourname@example.com"
TARGET_EMAIL = "your-email@example.com"

def get_all_emails(outlook):
    namespace = outlook.GetNamespace("MAPI")
    inbox = None

    # Method 1: match by account SMTP address, then find matching root folder
    target = TARGET_EMAIL.lower()
    for account in namespace.Accounts:
        if account.SmtpAddress.lower() == target:
            # Walk root folders to find one matching the account display name or email
            for folder in namespace.Folders:
                fname = folder.Name.lower()
                if account.DisplayName.lower() in fname or target in fname:
                    try:
                        inbox = folder.Folders["Inbox"]
                        break
                    except Exception:
                        pass
            break

    # Method 2: fallback - scan all root folders for one whose name contains the email or username
    if inbox is None:
        username = target.split("@")[0]
        for folder in namespace.Folders:
            fname = folder.Name.lower()
            if target in fname or username in fname:
                try:
                    inbox = folder.Folders["Inbox"]
                    break
                except Exception:
                    pass

    if inbox is None:
        # Debug: list all root folders so user can identify the right one
        print("ERROR: Could not find inbox for", TARGET_EMAIL)
        print("Available root folders in Outlook:")
        for folder in namespace.Folders:
            print(f"  - {folder.Name}")
        sys.exit("Update TARGET_EMAIL or check folder names above.")

    print(f"Using inbox: {inbox.Parent.Name} / {inbox.Name}")
    items = []
    _collect_items(inbox, items)
    return items

def _collect_items(folder, items):
    for item in folder.Items:
        try:
            if item.Class == 43:
                items.append(item)
        except Exception:
            pass
    for sub in folder.Folders:
        _collect_items(sub, items)

def is_rejection_email(mail):
    try:
        body    = (mail.Body    or "").lower()
        subject = (mail.Subject or "").lower()
        combined = body + " " + subject
        return any(phrase in combined for phrase in REJECTION_PHRASES)
    except Exception:
        return False

def is_interview_email(mail):
    try:
        body    = (mail.Body    or "").lower()
        subject = (mail.Subject or "").lower()
        combined = body + " " + subject
        return any(phrase in combined for phrase in INTERVIEW_PHRASES)
    except Exception:
        return False

def matches_company(mail, company):
    company_lower = company.lower().strip()
    # Strip legal suffixes for a looser core match
    company_core = re.sub(
        r"\b(inc|ltd|llc|gmbh|oy|ab|as|nv|plc|group|technologies|consulting|systems)\b",
        "", company_lower).strip()
    if len(company_core) < 3:
        company_core = company_lower

    try:
        sender_name  = (mail.SenderName         or "").lower()
        sender_email = (mail.SenderEmailAddress or "").lower()
        subject      = (mail.Subject            or "").lower()
        body         = (mail.Body               or "").lower()
    except Exception:
        return False

    domain_match = re.search(r"@([\w.-]+)", sender_email)
    domain      = domain_match.group(1) if domain_match else ""
    domain_root = domain.split(".")[0]

    # Word-boundary pattern for subject/body search
    pattern = re.compile(r"\b" + re.escape(company_core) + r"\b")

    # Match by sender fields OR company name anywhere in subject/body
    # This handles both direct company emails and ATS platforms naturally
    return (
        company_core in sender_name
        or company_lower in sender_name
        or company_core in sender_email
        or company_lower in domain
        or (len(company_core) >= 4 and company_core in domain_root)
        or bool(pattern.search(subject))
        or bool(pattern.search(body))
    )


ATS_DOMAINS = {
    "ashbyhq", "greenhouse", "lever", "workday", "taleo", "jobvite", "icims",
    "smartrecruiters", "bamboohr", "recruitee", "teamtailor", "workable",
    "myworkdayjobs", "imocha", "linkedin", "indeed", "glassdoor",
}

def is_application_email(mail):
    try:
        if is_rejection_email(mail) or is_interview_email(mail):
            return False
        body     = (mail.Body    or "").lower()
        subject  = (mail.Subject or "").lower()
        combined = body + " " + subject
        return any(phrase in combined for phrase in APPLICATION_PHRASES)
    except Exception:
        return False

def extract_company(mail):
    try:
        sender_name  = (mail.SenderName         or "").strip()
        sender_email = (mail.SenderEmailAddress or "").strip()
        subject      = (mail.Subject            or "")
        body         = (mail.Body               or "")
    except Exception:
        return None

    # 1. Clean sender name — strip generic HR/ATS noise words
    noise   = re.compile(
        r'\b(team|recruitment|recruiting|hiring|hr|careers|jobs|noreply|no.reply|'
        r'talent|people|notifications?|alerts?|updates?|info|donotreply|ops)\b',
        re.IGNORECASE,
    )
    cleaned = noise.sub("", sender_name).strip(" -|,.")
    # Accept if it's not just a person's name (First Last) and long enough
    if len(cleaned) >= 3 and not re.fullmatch(r'[A-Z][a-z]+ [A-Z][a-z]+', cleaned):
        return cleaned

    # 2. Body: "Thank you for considering Reaktor!"
    for pat in [
        r"(?:thank you for considering|thank you for your interest in)\s+([A-Z][A-Za-z0-9\s&]+?)(?:\s*[!.,])",
        r"joining\s+(?:the\s+)?([A-Z][A-Za-z0-9\s&]+?)\s+team",
    ]:
        m = re.search(pat, body)
        if m:
            company = m.group(1).strip()
            if 2 <= len(company) <= 50:
                return company

    # 3. Subject: "Welcome to Kuva Space", "application at Company"
    for pat in [
        r"welcome to\s+([A-Za-z0-9][A-Za-z0-9\s&]+?)(?:\s*[!.,]|$)",
        r"(?:at|from)\s+([A-Z][A-Za-z0-9\s&]+?)(?:\s*[!.,\-]|$)",
    ]:
        m = re.search(pat, subject, re.IGNORECASE)
        if m:
            company = m.group(1).strip()
            if 2 <= len(company) <= 50:
                return company

    # 4. Fallback: domain root if not an ATS platform
    m = re.search(r"@([\w.-]+)", sender_email)
    if m:
        domain_root = m.group(1).lower().split(".")[0]
        if domain_root not in ATS_DOMAINS and len(domain_root) >= 3:
            return domain_root.capitalize()

    return None

def extract_role(mail):
    try:
        subject = (mail.Subject or "")
        body    = (mail.Body    or "")
    except Exception:
        return None

    for pat in [
        r"(?:application (?:to|for)|applying (?:to|for))\s+(.+?)\s+(?:at|@)\s",
        r"(?:for|to)\s+(?:the\s+)?(.+?)\s+(?:position|role|opportunity)\b",
        r"[-–]\s*([A-Z][^-\n]{3,60}?)\s+(?:at\s+[A-Z]|\()",
    ]:
        m = re.search(pat, subject, re.IGNORECASE)
        if m:
            role = m.group(1).strip(" -|")
            if 3 <= len(role) <= 80:
                return role

    for pat in [
        r"(?:about our|for the|for our)\s+(.+?)\s+(?:position|role|opportunity)\b",
        r"your application(?:\s+\w+)?\s+(?:to|for)\s+(?:the\s+)?(.+?)(?:\s+(?:position|role|at)|[,\n])",
    ]:
        m = re.search(pat, body, re.IGNORECASE)
        if m:
            role = m.group(1).strip(" -|")
            if 3 <= len(role) <= 80:
                return role

    return None


# ---------- Excel helpers ----------

def load_wb(path):
    wb = openpyxl.load_workbook(path)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"ERROR: Sheet '{SHEET_NAME}' not found in {path}")
    return wb

def get_companies(ws):
    companies = {}
    for row in ws.iter_rows(min_row=DATA_START_ROW):
        cell = row[COL_COMPANY - 1]
        if cell.value:
            companies[cell.row] = str(cell.value).strip()
    return companies

def add_application_row(ws, company, role, date):
    last_row = DATA_START_ROW - 1
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_col=COL_COMPANY):
        if row[COL_COMPANY - 1].value is not None:
            last_row = row[COL_COMPANY - 1].row
    new_row = last_row + 1
    ws.cell(row=new_row, column=COL_COMPANY).value = company
    ws.cell(row=new_row, column=COL_ROLE).value    = role
    ws.cell(row=new_row, column=COL_DATE).value    = date
    ws.cell(row=new_row, column=COL_STATUS).value  = "Applied"
    ws.cell(row=new_row, column=COL_NOTES).value   = "Await response"

def reorder_and_format(ws):
    max_col = ws.max_column
    rows_data = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_col=max_col):
        if row[COL_COMPANY - 1].value is None:
            continue
        rows_data.append([cell.value for cell in row])

    def sort_key(r):
        return str(r[COL_COMPANY - 1] or "").lower()

    def status(r):
        return str(r[COL_STATUS - 1] or "").strip().lower()

    interviews = sorted([r for r in rows_data if status(r) == "interview"], key=sort_key)
    applied    = sorted([r for r in rows_data if status(r) == "applied"],   key=sort_key)
    declined   = sorted([r for r in rows_data if status(r) == "declined"],  key=sort_key)
    others     = sorted([r for r in rows_data if status(r) not in ("interview", "applied", "declined")], key=sort_key)

    font_interview = Font(color="00B050")  # green
    font_applied   = Font(color="BF8F00")  # dark yellow / gold
    font_declined  = Font(color="FF0000")  # red
    font_default   = Font()
    notes_col      = 4  # column E, zero-indexed

    status_fonts = {
        "interview": font_interview,
        "applied":   font_applied,
        "declined":  font_declined,
    }

    for i, row_vals in enumerate(interviews + applied + others + declined):
        row_num    = DATA_START_ROW + i
        row_status = status(row_vals)
        for j, val in enumerate(row_vals):
            cell = ws.cell(row=row_num, column=j + 1)
            if row_status == "declined" and j == notes_col and str(val or "").strip().lower() in ("await response", "awaiting for answer"):
                cell.value = None
            else:
                cell.value = val
            if j == COL_STATUS - 1:
                cell.font = status_fonts.get(row_status, font_default)


# ---------- Main ----------

def run(excel_path, dry_run, debug):
    print(f"\n{'[DRY RUN] ' if dry_run else ''}Job Tracker Auto-Decline")
    print(f"Excel    : {excel_path}")
    print(f"Phrases  : {len(REJECTION_PHRASES)} rejection, {len(INTERVIEW_PHRASES)} interview")
    print("-" * 55)

    wb = load_wb(excel_path)
    ws = wb[SHEET_NAME]
    companies = get_companies(ws)
    print(f"Found {len(companies)} companies in tracker.\n")

    print("Connecting to Outlook...")
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
    except Exception as e:
        sys.exit(f"ERROR: Cannot connect to Outlook.\n{e}")

    print("Fetching emails (this may take a moment)...")
    emails = get_all_emails(outlook)
    print(f"Loaded {len(emails)} emails from Inbox.\n")

    if debug:
        print("DEBUG - All emails:")
        for i, mail in enumerate(emails):
            try:
                combined = (mail.Body or "").lower() + " " + (mail.Subject or "").lower()
                rej_match = [p for p in REJECTION_PHRASES  if p in combined]
                int_match = [p for p in INTERVIEW_PHRASES  if p in combined]
                app_match = [p for p in APPLICATION_PHRASES if p in combined]
                flag = ""
                if rej_match: flag = f" <<< REJECTION: '{rej_match[0]}'"
                elif int_match: flag = f" <<< INTERVIEW: '{int_match[0]}'"
                elif app_match: flag = f" <<< APPLICATION: '{app_match[0]}'"
                print(f"  [{i+1:>3}] {(mail.SenderEmailAddress or ''):<40} | {(mail.Subject or '')[:55]}{flag}")
            except Exception:
                pass
        print()

    rejections = {}
    interviews = {}

    for row_num, company in companies.items():
        status_cell    = ws.cell(row=row_num, column=COL_STATUS)
        current_status = str(status_cell.value or "").strip().lower()

        for mail in emails:
            if not matches_company(mail, company):
                continue
            if current_status != "declined" and is_rejection_email(mail):
                rejections[row_num] = company
                if debug:
                    try: print(f"REJECTION MATCH: '{company}' <-- {mail.SenderEmailAddress} | {mail.Subject[:55]}")
                    except Exception: pass
                break
            if current_status == "applied" and is_interview_email(mail):
                interviews[row_num] = company
                if debug:
                    try: print(f"INTERVIEW MATCH: '{company}' <-- {mail.SenderEmailAddress} | {mail.Subject[:55]}")
                    except Exception: pass
                break

    # Detect new applications not yet in the tracker
    tracked_names = {str(c).lower() for c in companies.values()}
    new_apps = []
    for mail in emails:
        if not is_application_email(mail):
            continue
        company = extract_company(mail)
        if not company:
            continue
        c_low = company.lower().strip()
        if any(c_low in t or t in c_low for t in tracked_names):
            continue
        if c_low in {a["company"].lower() for a in new_apps}:
            continue
        role = extract_role(mail)
        try:
            rt   = mail.ReceivedTime
            date = datetime(rt.year, rt.month, rt.day)
        except Exception:
            date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        new_apps.append({"company": company, "role": role, "date": date})
        tracked_names.add(c_low)
        if debug:
            print(f"NEW APP: '{company}' | role: {role} | {date.strftime('%Y-%m-%d')}")

    if not rejections and not interviews and not new_apps:
        print("No updates found. Tracker is up to date.")
        if not debug:
            print("\nTip: run with --debug to see all emails and matched phrases.")
        return

    all_updates = {**interviews, **rejections}
    if all_updates:
        print(f"\n{'Would update' if dry_run else 'Updating'} {len(all_updates)} row(s):\n")
        for row_num, company in all_updates.items():
            status_cell = ws.cell(row=row_num, column=COL_STATUS)
            old_status  = status_cell.value
            new_status  = "Declined" if row_num in rejections else "Interview"
            print(f"  Row {row_num:>3}  {company:<35}  {old_status} -> {new_status}")
            if not dry_run:
                status_cell.value = new_status

    if new_apps:
        print(f"\n{'Would add' if dry_run else 'Adding'} {len(new_apps)} new application(s):\n")
        for app in new_apps:
            role_str = app["role"] or "role not found in email"
            print(f"  NEW   {app['company']:<35}  {role_str}")
            if not dry_run:
                add_application_row(ws, app["company"], app["role"], app["date"])

    if not dry_run:
        reorder_and_format(ws)
        wb.save(excel_path)
        print(f"\nSaved to {excel_path}  [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
    else:
        print("\n[DRY RUN] No changes saved. Remove --dry-run to apply.")


def main():
    parser = argparse.ArgumentParser(description="Auto-decline job apps based on Outlook rejection emails.")
    parser.add_argument("--excel",   type=Path, default=DEFAULT_EXCEL, help="Path to job_tracker.xlsx")
    parser.add_argument("--dry-run", action="store_true", help="Preview without saving")
    parser.add_argument("--debug",   action="store_true", help="Print all emails to diagnose missed matches")
    args = parser.parse_args()
    if not args.excel.exists():
        sys.exit(f"ERROR: Excel file not found: {args.excel}")
    run(args.excel, args.dry_run, args.debug)

if __name__ == "__main__":
    main()
