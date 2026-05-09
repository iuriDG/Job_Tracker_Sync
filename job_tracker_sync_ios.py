import argparse
import re
import sys
import imaplib
import email
import email.header
import email.utils
import getpass
import html
from pathlib import Path
from datetime import datetime, timezone, timedelta
from dataclasses import dataclass
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    sys.exit("ERROR: openpyxl is not installed.\nRun:  pip install openpyxl")


# ---------- Configuration ----------

DEFAULT_EXCEL = Path(__file__).parent / "job_tracker.xlsx"
COL_COMPANY   = 1   # Column A
COL_STATUS    = 4   # Column D

# Common IMAP servers — set the one matching your email provider:
#   Gmail:          imap.gmail.com        (requires App Password if 2FA is on)
#   Outlook/Hotmail: imap-mail.outlook.com
#   Yahoo:          imap.mail.yahoo.com   (requires App Password)
#   iCloud:         imap.mail.me.com      (requires App-Specific Password)
IMAP_SERVER = "imap.gmail.com"
IMAP_PORT   = 993

# Set to limit how far back to scan (0 = all emails)
DAYS_BACK = 180

# --- Rejection phrases ---
REJECTION_PHRASES = [
    "we will not be moving forward",
    "we won't be moving forward",
    "we are not moving forward",
    "we're not moving forward",
    "not moving forward with your application",
    "not moving forward with you",
    "decided not to move forward",
    "have decided not to move forward",
    "unable to move forward",
    "we will not be proceeding",
    "we won't be proceeding",
    "we are not proceeding",
    "not proceeding with your application",
    "not proceeding with your candidacy",
    "unable to proceed with your candidacy",
    "unable to proceed with your application",
    "we're unable to proceed",
    "we are unable to proceed",
    "we have decided to move forward with other candidates",
    "decided to move forward with other candidates",
    "moving forward with other candidates",
    "moving forward with another candidate",
    "we have chosen to move forward with another candidate",
    "we have chosen to move forward with other candidates",
    "we have chosen other candidates",
    "chosen to pursue other candidates",
    "we will be moving forward with other applicants",
    "pursuing other candidates",
    "pursuing other applicants",
    "pursuing other profiles",
    "you have not been selected",
    "not been selected for",
    "not selected for this position",
    "not selected for this role",
    "not selected for the position",
    "not selected for the role",
    "not selected to move forward",
    "was not selected",
    "were not selected",
    "your application was not successful",
    "your application has been unsuccessful",
    "unsuccessful in your application",
    "we are unable to offer you",
    "we will not be offering you",
    "we won't be offering you",
    "no offer will be extended",
    "we will not be extending an offer",
    "we regret to inform you",
    "we regret to let you know",
    "we regret to advise you",
    "we're sorry to inform you",
    "we are sorry to inform you",
    "we're sorry to let you know",
    "we are sorry to let you know",
    "we're sorry to advise you",
    "sorry to let you know",
    "sorry to inform you",
    "after careful consideration",
    "after thorough consideration",
    "after reviewing your application",
    "after much consideration",
    "following careful consideration",
    "following our review",
    "having reviewed your application",
    "position has been filled",
    "role has been filled",
    "vacancy has been filled",
    "the position is no longer available",
    "we have filled the position",
    "we have closed this position",
    "does not meet our requirements",
    "do not meet the requirements",
    "does not match our current needs",
    "not the right fit",
    "not a strong enough match",
    "not quite what we are looking for",
    "not quite what we're looking for",
    "your profile does not match",
    "your background does not match",
    "your experience does not match",
    "thank you for your interest, however",
    "thank you for applying, however",
    "appreciate your interest, however",
    "we have decided not to",
    "decided to decline",
    "we will not be considering",
    "no longer being considered",
    "your application will not be taken further",
    "will not be taken further",
    "will not be progressed",
    "your candidacy will not",
]

# --- Interview / invitation phrases ---
INTERVIEW_PHRASES = [
    "invite you to an interview",
    "invited you to an interview",
    "invitation to interview",
    "invitation for an interview",
    "you are invited to interview",
    "you have been invited to interview",
    "we would like to invite you to interview",
    "pleased to invite you to an interview",
    "happy to invite you to an interview",
    "we would like to interview you",
    "we'd like to interview you",
    "like to interview you",
    "we want to interview you",
    "you have been selected for an interview",
    "you've been selected for an interview",
    "selected to move forward to an interview",
    "shortlisted for an interview",
    "you have been shortlisted",
    "you've been shortlisted",
    "moving forward with you to the interview",
    "moving you forward to the next stage",
    "moving you to the next round",
    "you have been selected for the next stage",
    "selected for the next round",
    "you have been selected to proceed",
    "you've been selected to proceed",
    "schedule an interview",
    "schedule a interview",
    "schedule your interview",
    "scheduling an interview",
    "like to schedule a time",
    "like to set up a time",
    "like to set up a call",
    "like to schedule a call",
    "like to arrange a call",
    "like to arrange an interview",
    "like to arrange a meeting",
    "schedule a meeting with you",
    "schedule time to talk",
    "schedule a conversation",
    "schedule a chat",
    "book an interview",
    "book a time with",
    "book a call",
    "book a meeting",
    "next step is an interview",
    "next steps involve an interview",
    "next step in our process",
    "next stage of our process",
    "next stage of the recruitment",
    "next stage of the hiring",
    "move you to the next step",
    "progress to the next stage",
    "progress to the interview stage",
    "advance to the next round",
    "advance you to the interview",
    "phone screen",
    "phone interview",
    "video interview",
    "video call with",
    "virtual interview",
    "zoom interview",
    "teams interview",
    "google meet interview",
    "introductory call",
    "initial call",
    "screening call",
    "discovery call",
    "chat with our team",
    "chat with the team",
    "speak with our recruiter",
    "speak with the hiring manager",
    "talk with our team",
    "technical assessment",
    "technical interview",
    "technical test",
    "coding assessment",
    "coding challenge",
    "coding test",
    "take-home assignment",
    "take home assignment",
    "complete an assessment",
    "complete a test",
    "online assessment",
    "would like to meet with you",
    "we would like to invite you",
    "we'd like to invite you",
    "we'd love to meet you",
    "would love to meet you",
    "would love to chat with you",
    "would love to connect",
    "pleased to meet with you",
    "happy to meet with you",
    "keen to learn more about you",
    "excited to learn more about you",
    "we came across your profile",
    "we found your profile",
    "your profile caught our attention",
    "your background caught our attention",
    "your experience caught our attention",
    "i came across your profile",
    "i found your profile",
    "reach out regarding an opportunity",
    "reaching out about an opportunity",
    "reaching out about a role",
    "reaching out about a position",
    "open to exploring",
    "open to new opportunities",
    "interested in hearing more",
]

# --- Application confirmation phrases ---
APPLICATION_PHRASES = [
    "thank you for your application",
    "thank you for applying",
    "we have received your application",
    "we received your application",
    "your application has been received",
    "application has been submitted",
    "thank you for considering",
    "we are delighted that you",
    "we're delighted that you",
    "thanks for applying",
    "your application to",
    "we will be in touch",
    "we'll be in touch",
    "will get back to you",
    "will be in touch as soon as",
]

SHEET_NAME     = "Job Tracker"
DATA_START_ROW = 5
COL_ROLE  = 2   # Column B
COL_DATE  = 3   # Column C
COL_NOTES = 5   # Column E

ATS_DOMAINS = {
    "ashbyhq", "greenhouse", "lever", "workday", "taleo", "jobvite", "icims",
    "smartrecruiters", "bamboohr", "recruitee", "teamtailor", "workable",
    "myworkdayjobs", "imocha", "linkedin", "indeed", "glassdoor",
}


# ---------- Mail item wrapper ----------

@dataclass
class MailItem:
    Subject: str
    Body: str
    SenderName: str
    SenderEmailAddress: str
    ReceivedTime: datetime


# ---------- IMAP helpers ----------

def _decode_header(value: str) -> str:
    parts = email.header.decode_header(value or "")
    decoded = []
    for chunk, charset in parts:
        if isinstance(chunk, bytes):
            decoded.append(chunk.decode(charset or "utf-8", errors="replace"))
        else:
            decoded.append(chunk)
    return "".join(decoded)


def _extract_body(msg) -> str:
    """Return the best plain-text body from a (possibly multipart) message."""
    plain = ""
    html_body = ""
    if msg.is_multipart():
        for part in msg.walk():
            ct = part.get_content_type()
            cd = str(part.get("Content-Disposition") or "")
            if "attachment" in cd:
                continue
            charset = part.get_content_charset() or "utf-8"
            if ct == "text/plain" and not plain:
                plain = part.get_payload(decode=True).decode(charset, errors="replace")
            elif ct == "text/html" and not html_body:
                html_body = part.get_payload(decode=True).decode(charset, errors="replace")
    else:
        charset = msg.get_content_charset() or "utf-8"
        payload = msg.get_payload(decode=True) or b""
        if msg.get_content_type() == "text/html":
            html_body = payload.decode(charset, errors="replace")
        else:
            plain = payload.decode(charset, errors="replace")

    if plain:
        return plain

    # Strip HTML tags as a fallback
    text = re.sub(r"<[^>]+>", " ", html_body)
    return html.unescape(text)


def _parse_message(raw: bytes) -> Optional[MailItem]:
    try:
        msg = email.message_from_bytes(raw)

        subject = _decode_header(msg.get("Subject", ""))
        from_raw = _decode_header(msg.get("From", ""))
        name, addr = email.utils.parseaddr(from_raw)
        body = _extract_body(msg)

        date_str = msg.get("Date", "")
        try:
            received = email.utils.parsedate_to_datetime(date_str)
            # Normalize to naive local datetime
            if received.tzinfo is not None:
                received = received.astimezone().replace(tzinfo=None)
        except Exception:
            received = datetime.now()

        return MailItem(
            Subject=subject,
            Body=body,
            SenderName=name,
            SenderEmailAddress=addr,
            ReceivedTime=received,
        )
    except Exception:
        return None


def get_all_emails(email_address: str, password: str) -> list:
    print(f"Connecting to {IMAP_SERVER}:{IMAP_PORT} ...")
    try:
        conn = imaplib.IMAP4_SSL(IMAP_SERVER, IMAP_PORT)
        conn.login(email_address, password)
    except imaplib.IMAP4.error as e:
        sys.exit(f"ERROR: IMAP login failed.\n{e}\n\n"
                 "If using Gmail with 2FA, generate an App Password at:\n"
                 "  https://myaccount.google.com/apppasswords")

    conn.select("INBOX", readonly=True)

    if DAYS_BACK > 0:
        since = (datetime.now() - timedelta(days=DAYS_BACK)).strftime("%d-%b-%Y")
        status, data = conn.search(None, f'SINCE "{since}"')
    else:
        status, data = conn.search(None, "ALL")

    if status != "OK":
        sys.exit("ERROR: Could not search inbox.")

    msg_ids = data[0].split()
    print(f"Fetching {len(msg_ids)} emails from Inbox (last {DAYS_BACK} days)...")

    items = []
    for i, uid in enumerate(msg_ids, 1):
        if i % 200 == 0:
            print(f"  ... {i}/{len(msg_ids)}")
        status, raw = conn.fetch(uid, "(RFC822)")
        if status != "OK" or not raw or not raw[0]:
            continue
        item = _parse_message(raw[0][1])
        if item:
            items.append(item)

    conn.logout()
    return items


# ---------- Email classification helpers ----------

def is_rejection_email(mail: MailItem) -> bool:
    combined = (mail.Body + " " + mail.Subject).lower()
    return any(phrase in combined for phrase in REJECTION_PHRASES)


def is_interview_email(mail: MailItem) -> bool:
    combined = (mail.Body + " " + mail.Subject).lower()
    return any(phrase in combined for phrase in INTERVIEW_PHRASES)


def is_application_email(mail: MailItem) -> bool:
    if is_rejection_email(mail) or is_interview_email(mail):
        return False
    combined = (mail.Body + " " + mail.Subject).lower()
    return any(phrase in combined for phrase in APPLICATION_PHRASES)


def matches_company(mail: MailItem, company: str) -> bool:
    company_lower = company.lower().strip()
    company_core = re.sub(
        r"\b(inc|ltd|llc|gmbh|oy|ab|as|nv|plc|group|technologies|consulting|systems)\b",
        "", company_lower).strip()
    if len(company_core) < 3:
        company_core = company_lower

    sender_name  = mail.SenderName.lower()
    sender_email = mail.SenderEmailAddress.lower()
    subject      = mail.Subject.lower()
    body         = mail.Body.lower()

    domain_match = re.search(r"@([\w.-]+)", sender_email)
    domain      = domain_match.group(1) if domain_match else ""
    domain_root = domain.split(".")[0]

    pattern = re.compile(r"\b" + re.escape(company_core) + r"\b")

    return (
        company_core in sender_name
        or company_lower in sender_name
        or company_core in sender_email
        or company_lower in domain
        or (len(company_core) >= 4 and company_core in domain_root)
        or bool(pattern.search(subject))
        or bool(pattern.search(body))
    )


def extract_company(mail: MailItem) -> Optional[str]:
    sender_name  = mail.SenderName.strip()
    sender_email = mail.SenderEmailAddress.strip()
    subject      = mail.Subject
    body         = mail.Body

    noise = re.compile(
        r'\b(team|recruitment|recruiting|hiring|hr|careers|jobs|noreply|no.reply|'
        r'talent|people|notifications?|alerts?|updates?|info|donotreply|ops)\b',
        re.IGNORECASE,
    )
    cleaned = noise.sub("", sender_name).strip(" -|,.")
    if len(cleaned) >= 3 and not re.fullmatch(r'[A-Z][a-z]+ [A-Z][a-z]+', cleaned):
        return cleaned

    for pat in [
        r"(?:thank you for considering|thank you for your interest in)\s+([A-Z][A-Za-z0-9\s&]+?)(?:\s*[!.,])",
        r"joining\s+(?:the\s+)?([A-Z][A-Za-z0-9\s&]+?)\s+team",
    ]:
        m = re.search(pat, body)
        if m:
            company = m.group(1).strip()
            if 2 <= len(company) <= 50:
                return company

    for pat in [
        r"welcome to\s+([A-Za-z0-9][A-Za-z0-9\s&]+?)(?:\s*[!.,]|$)",
        r"(?:at|from)\s+([A-Z][A-Za-z0-9\s&]+?)(?:\s*[!.,\-]|$)",
    ]:
        m = re.search(pat, subject, re.IGNORECASE)
        if m:
            company = m.group(1).strip()
            if 2 <= len(company) <= 50:
                return company

    m = re.search(r"@([\w.-]+)", sender_email)
    if m:
        domain_root = m.group(1).lower().split(".")[0]
        if domain_root not in ATS_DOMAINS and len(domain_root) >= 3:
            return domain_root.capitalize()

    return None


def extract_role(mail: MailItem) -> Optional[str]:
    subject = mail.Subject
    body    = mail.Body

    for pat in [
        r"(?:application (?:to|for)|applying (?:to|for))\s+(.+?)\s+(?:at|@)\s",
        r"(?:for|to)\s+(?:the\s+)?(.+?)\s+(?:position|role|opportunity)\b",
        r"[-–]\s*([A-Z][^-\n]{3,60}?)\s+(?:at\s+[A-Z]|\()",
    ]:
        m = re.search(pat, subject, re.IGNORECASE)
        if m:
            role = m.group(1).strip(" -|")
            if 3 <= len(role) <= 80:
                return role

    for pat in [
        r"(?:about our|for the|for our)\s+(.+?)\s+(?:position|role|opportunity)\b",
        r"your application(?:\s+\w+)?\s+(?:to|for)\s+(?:the\s+)?(.+?)(?:\s+(?:position|role|at)|[,\n])",
    ]:
        m = re.search(pat, body, re.IGNORECASE)
        if m:
            role = m.group(1).strip(" -|")
            if 3 <= len(role) <= 80:
                return role

    return None


# ---------- Excel helpers ----------

def load_wb(path: Path):
    wb = openpyxl.load_workbook(path)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"ERROR: Sheet '{SHEET_NAME}' not found in {path}")
    return wb


def get_companies(ws) -> dict:
    companies = {}
    for row in ws.iter_rows(min_row=DATA_START_ROW):
        cell = row[COL_COMPANY - 1]
        if cell.value:
            companies[cell.row] = str(cell.value).strip()
    return companies


def add_application_row(ws, company: str, role: Optional[str], date: datetime):
    last_row = DATA_START_ROW - 1
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_col=COL_COMPANY):
        if row[COL_COMPANY - 1].value is not None:
            last_row = row[COL_COMPANY - 1].row
    new_row = last_row + 1
    ws.cell(row=new_row, column=COL_COMPANY).value = company
    ws.cell(row=new_row, column=COL_ROLE).value    = role
    ws.cell(row=new_row, column=COL_DATE).value    = date
    ws.cell(row=new_row, column=COL_STATUS).value  = "Applied"
    ws.cell(row=new_row, column=COL_NOTES).value   = "Await response"


def reorder_and_format(ws):
    max_col   = ws.max_column
    rows_data = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_col=max_col):
        if row[COL_COMPANY - 1].value is None:
            continue
        rows_data.append([cell.value for cell in row])

    def sort_key(r):
        return str(r[COL_COMPANY - 1] or "").lower()

    def status(r):
        return str(r[COL_STATUS - 1] or "").strip().lower()

    interviews = sorted([r for r in rows_data if status(r) == "interview"], key=sort_key)
    applied    = sorted([r for r in rows_data if status(r) == "applied"],   key=sort_key)
    declined   = sorted([r for r in rows_data if status(r) == "declined"],  key=sort_key)
    others     = sorted([r for r in rows_data if status(r) not in ("interview", "applied", "declined")], key=sort_key)

    font_interview = Font(color="00B050")
    font_applied   = Font(color="BF8F00")
    font_declined  = Font(color="FF0000")
    font_default   = Font()
    notes_col      = 4  # Column E, zero-indexed

    status_fonts = {
        "interview": font_interview,
        "applied":   font_applied,
        "declined":  font_declined,
    }

    for i, row_vals in enumerate(interviews + applied + others + declined):
        row_num    = DATA_START_ROW + i
        row_status = status(row_vals)
        for j, val in enumerate(row_vals):
            cell = ws.cell(row=row_num, column=j + 1)
            if row_status == "declined" and j == notes_col and str(val or "").strip().lower() in ("await response", "awaiting for answer"):
                cell.value = None
            else:
                cell.value = val
            if j == COL_STATUS - 1:
                cell.font = status_fonts.get(row_status, font_default)


# ---------- Main ----------

def run(excel_path: Path, email_address: str, password: str, dry_run: bool, debug: bool):
    print(f"\n{'[DRY RUN] ' if dry_run else ''}Job Tracker Auto-Sync (macOS / IMAP)")
    print(f"Excel    : {excel_path}")
    print(f"Account  : {email_address}")
    print(f"Phrases  : {len(REJECTION_PHRASES)} rejection, {len(INTERVIEW_PHRASES)} interview")
    print("-" * 55)

    wb       = load_wb(excel_path)
    ws       = wb[SHEET_NAME]
    companies = get_companies(ws)
    print(f"Found {len(companies)} companies in tracker.\n")

    emails = get_all_emails(email_address, password)
    print(f"Loaded {len(emails)} emails.\n")

    if debug:
        print("DEBUG - All emails:")
        for i, mail in enumerate(emails):
            combined   = (mail.Body + " " + mail.Subject).lower()
            rej_match  = [p for p in REJECTION_PHRASES   if p in combined]
            int_match  = [p for p in INTERVIEW_PHRASES   if p in combined]
            app_match  = [p for p in APPLICATION_PHRASES if p in combined]
            flag = ""
            if rej_match:  flag = f" <<< REJECTION: '{rej_match[0]}'"
            elif int_match: flag = f" <<< INTERVIEW: '{int_match[0]}'"
            elif app_match: flag = f" <<< APPLICATION: '{app_match[0]}'"
            print(f"  [{i+1:>3}] {mail.SenderEmailAddress:<40} | {mail.Subject[:55]}{flag}")
        print()

    rejections: dict = {}
    interviews: dict = {}

    for row_num, company in companies.items():
        status_cell    = ws.cell(row=row_num, column=COL_STATUS)
        current_status = str(status_cell.value or "").strip().lower()

        for mail in emails:
            if not matches_company(mail, company):
                continue
            if current_status != "declined" and is_rejection_email(mail):
                rejections[row_num] = company
                if debug:
                    print(f"REJECTION MATCH: '{company}' <-- {mail.SenderEmailAddress} | {mail.Subject[:55]}")
                break
            if current_status == "applied" and is_interview_email(mail):
                interviews[row_num] = company
                if debug:
                    print(f"INTERVIEW MATCH: '{company}' <-- {mail.SenderEmailAddress} | {mail.Subject[:55]}")
                break

    tracked_names = {str(c).lower() for c in companies.values()}
    new_apps = []
    for mail in emails:
        if not is_application_email(mail):
            continue
        company = extract_company(mail)
        if not company:
            continue
        c_low = company.lower().strip()
        if any(c_low in t or t in c_low for t in tracked_names):
            continue
        if c_low in {a["company"].lower() for a in new_apps}:
            continue
        role = extract_role(mail)
        rt   = mail.ReceivedTime
        date = datetime(rt.year, rt.month, rt.day)
        new_apps.append({"company": company, "role": role, "date": date})
        tracked_names.add(c_low)
        if debug:
            print(f"NEW APP: '{company}' | role: {role} | {date.strftime('%Y-%m-%d')}")

    if not rejections and not interviews and not new_apps:
        print("No updates found. Tracker is up to date.")
        if not debug:
            print("\nTip: run with --debug to see all emails and matched phrases.")
        return

    all_updates = {**interviews, **rejections}
    if all_updates:
        print(f"\n{'Would update' if dry_run else 'Updating'} {len(all_updates)} row(s):\n")
        for row_num, company in all_updates.items():
            status_cell = ws.cell(row=row_num, column=COL_STATUS)
            old_status  = status_cell.value
            new_status  = "Declined" if row_num in rejections else "Interview"
            print(f"  Row {row_num:>3}  {company:<35}  {old_status} -> {new_status}")
            if not dry_run:
                status_cell.value = new_status

    if new_apps:
        print(f"\n{'Would add' if dry_run else 'Adding'} {len(new_apps)} new application(s):\n")
        for app in new_apps:
            role_str = app["role"] or "role not found in email"
            print(f"  NEW   {app['company']:<35}  {role_str}")
            if not dry_run:
                add_application_row(ws, app["company"], app["role"], app["date"])

    if not dry_run:
        reorder_and_format(ws)
        wb.save(excel_path)
        print(f"\nSaved to {excel_path}  [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
    else:
        print("\n[DRY RUN] No changes saved. Remove --dry-run to apply.")


def main():
    parser = argparse.ArgumentParser(
        description="Sync job application statuses from your email inbox to Excel tracker (macOS / IMAP)."
    )
    parser.add_argument("--excel",   type=Path, default=DEFAULT_EXCEL, help="Path to job_tracker.xlsx")
    parser.add_argument("--email",   type=str,  required=True,         help="Your email address")
    parser.add_argument("--password",type=str,  default=None,          help="Email password or app password (prompted if omitted)")
    parser.add_argument("--server",  type=str,  default=IMAP_SERVER,   help=f"IMAP server (default: {IMAP_SERVER})")
    parser.add_argument("--port",    type=int,  default=IMAP_PORT,     help=f"IMAP port (default: {IMAP_PORT})")
    parser.add_argument("--days",    type=int,  default=DAYS_BACK,     help="How many days back to scan (0 = all)")
    parser.add_argument("--dry-run", action="store_true",              help="Preview without saving")
    parser.add_argument("--debug",   action="store_true",              help="Print all emails to diagnose missed matches")
    args = parser.parse_args()

    if not args.excel.exists():
        sys.exit(f"ERROR: Excel file not found: {args.excel}")

    # Override module-level config from CLI args
    global IMAP_SERVER, IMAP_PORT, DAYS_BACK
    IMAP_SERVER = args.server
    IMAP_PORT   = args.port
    DAYS_BACK   = args.days

    password = args.password or getpass.getpass(f"Password / App Password for {args.email}: ")

    run(args.excel, args.email, password, args.dry_run, args.debug)


if __name__ == "__main__":
    main()
